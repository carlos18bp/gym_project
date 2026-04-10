"""
Document-related Excel report generators.
"""
import pandas as pd
from django.db import models
from django.utils import timezone
from rest_framework.response import Response
from rest_framework import status
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint

from gym_app.models import DynamicDocument, User


def _get_user_id():
    """Read user_id from the package so tests can patch gym_app.views.reports.user_id."""
    import gym_app.views.reports as pkg
    return pkg.user_id


def generate_documents_by_state_report(response, start_date, end_datetime):
    """
    Generate a report of documents by state with filtering options.
    """
    # Define colors for document states
    state_colors = {
        'Draft': 'FFFFE0',      # Light yellow
        'Published': 'E6F9FF',  # Light blue
        'Progress': 'E6FFE6',   # Light green
        'Completed': 'E6E6E6',  # Light gray
    }

    # State translations for the report
    state_translations = {
        'Draft': 'Borrador',
        'Published': 'Publicado',
        'Progress': 'En Progreso',
        'Completed': 'Completado',
    }

    # Query documents within the date range
    documents_query = DynamicDocument.objects.filter(
        models.Q(created_at__date__gte=start_date, created_at__date__lte=end_datetime.date()) |
        models.Q(updated_at__date__gte=start_date, updated_at__date__lte=end_datetime.date())
    ).select_related('created_by', 'assigned_to')

    # Apply user filter if specified
    user_id = _get_user_id()
    if user_id:
        try:
            user = User.objects.get(id=user_id)
            if user.role == 'lawyer':
                documents_query = documents_query.filter(
                    models.Q(created_by=user) | models.Q(assigned_to=user)
                )
            elif user.role == 'client':
                documents_query = documents_query.filter(assigned_to=user)
        except User.DoesNotExist:
            return Response(
                {'error': f'User with id {user_id} does not exist'},
                status=status.HTTP_404_NOT_FOUND
            )

    # Get the document data
    document_data = []
    today = timezone.now().date()

    for doc in documents_query:
        days_since_update = (today - doc.updated_at.date()).days

        creator_name = f"{doc.created_by.first_name} {doc.created_by.last_name}" if doc.created_by else "N/A"
        assignee_name = f"{doc.assigned_to.first_name} {doc.assigned_to.last_name}" if doc.assigned_to else "N/A"

        document_data.append({
            'title': doc.title,
            'state': doc.state,
            'state_es': state_translations.get(doc.state, doc.state),
            'created_by': creator_name,
            'assigned_to': assignee_name,
            'created_at': doc.created_at,
            'updated_at': doc.updated_at,
            'days_since_update': days_since_update
        })

    # Create dataframe for Excel export
    if document_data:
        df = pd.DataFrame(document_data)

        # Format dates
        df['created_at'] = df['created_at'].dt.strftime('%Y-%m-%d %H:%M')
        df['updated_at'] = df['updated_at'].dt.strftime('%Y-%m-%d %H:%M')

        # Create state counts for summary
        state_counts = df['state'].value_counts().to_dict()

        # Create Excel writer
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Documentos por Estado', index=False, columns=[
                'title', 'state_es', 'created_by', 'assigned_to',
                'created_at', 'updated_at', 'days_since_update'
            ])

            # Get the worksheet to apply formatting
            worksheet = writer.sheets['Documentos por Estado']

            # Set column names in Spanish
            column_names = [
                'Título', 'Estado', 'Creado por', 'Asignado a',
                'Fecha Creación', 'Última Actualización', 'Días desde Actualización'
            ]

            for idx, name in enumerate(column_names, start=1):
                cell = worksheet.cell(row=1, column=idx)
                cell.value = name
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")

            # Apply conditional formatting based on state
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for row_idx, row in enumerate(df.itertuples(), start=2):
                state = row.state
                color = state_colors.get(state, 'FFFFFF')  # Default white if state not found

                for col_idx in range(1, len(column_names) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')

                    # Center certain columns
                    if col_idx in [2, 5, 6, 7]:  # State, dates, and days columns
                        cell.alignment = Alignment(horizontal='center', vertical='center')

            # Auto-adjust column widths
            columns_to_export = ['title', 'state_es', 'created_by', 'assigned_to', 'created_at', 'updated_at', 'days_since_update']
            for idx, col in enumerate(columns_to_export):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(column_names[idx])
                ) + 2  # padding
                worksheet.column_dimensions[get_column_letter(idx+1)].width = max_length

            # Add summary sheet
            summary_data = {
                'Estado': [state_translations.get(state, state) for state in state_counts.keys()],
                'Cantidad': list(state_counts.values())
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Resumen', index=False)

            # Format summary sheet
            summary_sheet = writer.sheets['Resumen']
            summary_sheet.column_dimensions['A'].width = 15
            summary_sheet.column_dimensions['B'].width = 10

            # Format header
            for col in ['A', 'B']:
                cell = summary_sheet[f'{col}1']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')

            # Apply colors to state rows
            for row_idx, state in enumerate(state_counts.keys(), start=2):
                color = state_colors.get(state, 'FFFFFF')
                cell_state = summary_sheet[f'A{row_idx}']
                cell_count = summary_sheet[f'B{row_idx}']

                for cell in [cell_state, cell_count]:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')

            # Add a total row
            total_row = len(state_counts) + 2
            summary_sheet[f'A{total_row}'] = 'Total'
            summary_sheet[f'B{total_row}'] = sum(state_counts.values())

            for col in ['A', 'B']:
                cell = summary_sheet[f'{col}{total_row}']
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            # Add a pie chart
            pie = PieChart()
            labels = Reference(summary_sheet, min_col=1, min_row=2, max_row=total_row-1)
            data = Reference(summary_sheet, min_col=2, min_row=1, max_row=total_row-1)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Distribución de Documentos por Estado"

            # Customize chart slices
            slice_colors = ['FFFFE0', 'E6F9FF', 'E6FFE6', 'E6E6E6']
            for i, color in enumerate(slice_colors[:len(state_counts)]):
                slice = DataPoint(idx=i)
                slice.graphicalProperties.solidFill = color
                pie.series[0].data_points.append(slice)

            # Add the chart to the sheet
            summary_sheet.add_chart(pie, "D2")

        return response
    else:
        # Return empty Excel if no data found
        df = pd.DataFrame({
            'Mensaje': ['No se encontraron documentos para el período seleccionado.']
        })
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sin Datos', index=False)
        return response
