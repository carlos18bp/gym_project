"""
Reports views for generating Excel reports.
"""
import io
import datetime
from datetime import timedelta
import pandas as pd
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, F, Q
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.utils import timezone
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.series import DataPoint

from gym_app.models import Process, Case, Stage, User, ActivityFeed, DynamicDocument, LegalRequest, LegalRequestType, LegalDiscipline


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def generate_excel_report(request):
    """
    Generate an Excel report based on the provided parameters.
    If dates are not provided, all data will be included without date filtering.
    """
    report_type = request.data.get('reportType')
    start_date = request.data.get('startDate')
    end_date = request.data.get('endDate')
    
    if not report_type:
        return Response(
            {'error': 'reportType is required'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Initialize dates to None for no filtering if not provided
    start_datetime = None
    end_datetime = None
    
    # Process dates if provided
    if start_date and end_date:
        try:
            # Convert string dates to datetime objects
            start_datetime = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
            end_datetime = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()
            
            # Set end_date to end of day for inclusive filtering
            end_datetime = datetime.datetime.combine(end_datetime, datetime.time.max)
        except ValueError:
            return Response(
                {'error': 'Invalid date format. Use YYYY-MM-DD'},
                status=status.HTTP_400_BAD_REQUEST
            )
    # If only one date is provided, require both
    elif start_date or end_date:
        return Response(
            {'error': 'Both startDate and endDate must be provided if using date filtering'},
            status=status.HTTP_400_BAD_REQUEST
        )
    # If no dates, use earliest possible date and today's end date
    else:
        start_datetime = datetime.datetime(1900, 1, 1)
        end_datetime = datetime.datetime.combine(timezone.now().date(), datetime.time.max)
    
    # Initialize response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{report_type}_{datetime.date.today()}.xlsx"'
    
    # Generate appropriate report based on type
    if report_type == 'active_processes':
        return generate_active_processes_report(response, start_datetime, end_datetime)
    elif report_type == 'processes_by_lawyer':
        return generate_processes_by_lawyer_report(response, start_datetime, end_datetime)
    elif report_type == 'processes_by_client':
        return generate_processes_by_client_report(response, start_datetime, end_datetime)
    elif report_type == 'process_stages':
        return generate_process_stages_report(response, start_datetime, end_datetime)
    elif report_type == 'registered_users':
        return generate_registered_users_report(response, start_datetime, end_datetime)
    elif report_type == 'user_activity':
        return generate_user_activity_report(response, start_datetime, end_datetime)
    elif report_type == 'lawyers_workload':
        return generate_lawyers_workload_report(response, start_datetime, end_datetime)
    elif report_type == 'documents_by_state':
        return generate_documents_by_state_report(response, start_datetime, end_datetime)
    elif report_type == 'received_legal_requests':
        return generate_received_legal_requests_report(response, start_datetime, end_datetime)
    elif report_type == 'requests_by_type_discipline':
        return generate_requests_by_type_discipline_report(response, start_datetime, end_datetime)
    else:
        return Response(
            {'error': f'Report type {report_type} not supported'},
            status=status.HTTP_400_BAD_REQUEST
        )


def generate_active_processes_report(response, start_date, end_datetime):
    """
    Generate report of active processes with detailed information.
    
    Columns:
    - Referencia (Process.ref)
    - Tipo de Caso (Case.type)
    - Subcaso (Process.subcase)
    - Cliente (User.first_name + User.last_name)
    - Abogado Asignado (User.first_name + User.last_name)
    - Autoridad (Process.authority)
    - Demandante (Process.plaintiff)
    - Demandado (Process.defendant)
    - Etapa Actual (Stage.status del último registro)
    - Fecha de Creación (Process.created_at)
    - Días Activo (calculado)
    """
    # Base queryset with appropriate filtering
    processes = Process.objects.filter(
        created_at__range=[start_date, end_datetime]
    ).prefetch_related(
        'stages', 'client', 'lawyer', 'case'
    )
    
    # Prepare data for Excel
    data = []
    today = datetime.date.today()
    
    for process in processes:
        # Get the latest stage (assuming stages are ordered by created_at)
        latest_stage = process.stages.order_by('-created_at').first()
        current_stage = latest_stage.status if latest_stage else "Sin etapa"
        
        # Calculate days active
        created_date = process.created_at.date()
        days_active = (today - created_date).days
        
        # Client and lawyer names
        client_name = f"{process.client.first_name or ''} {process.client.last_name or ''}".strip()
        lawyer_name = f"{process.lawyer.first_name or ''} {process.lawyer.last_name or ''}".strip()
        
        # Add process data to the list
        data.append({
            'Referencia': process.ref,
            'Tipo de Caso': process.case.type,
            'Subcaso': process.subcase,
            'Cliente': client_name,
            'Abogado Asignado': lawyer_name,
            'Autoridad': process.authority,
            'Demandante': process.plaintiff,
            'Demandado': process.defendant,
            'Etapa Actual': current_stage,
            'Fecha de Creación': created_date,
            'Días Activo': days_active
        })
    
    # Create DataFrame from data
    df = pd.DataFrame(data)
    
    # Create Excel file
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Procesos Activos', index=False)
        
        # Get workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Procesos Activos']
        
        # Add formats
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#D9D9D9',
            'border': 1
        })
        
        # Write headers with format
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            # Set column width based on max length in column
            if not df.empty:
                max_len = max(
                    df[value].astype(str).map(len).max(),
                    len(value)
                ) + 2
                worksheet.set_column(col_num, col_num, max_len)
    
    return response


def generate_processes_by_lawyer_report(response, start_date, end_datetime):
    """
    Generate report of processes grouped by lawyer.
    
    Columns:
    - Abogado (User.first_name + User.last_name + User.email)
    - Referencia de Proceso (Process.ref)
    - Tipo de Caso (Case.type)
    - Cliente (User.first_name + User.last_name)
    - Etapa Actual (Stage.status)
    - Fecha de Creación (Process.created_at)
    - Días Activo (calculado)
    """
    # Filter lawyers if a specific user_id is provided
    if user_id:
        lawyers = User.objects.filter(id=user_id, role='lawyer')
    else:
        lawyers = User.objects.filter(role='lawyer')
    
    # Prepare data for Excel
    data = []
    today = datetime.date.today()
    
    for lawyer in lawyers:
        # Get all processes for this lawyer
        processes = Process.objects.filter(
            lawyer=lawyer,
            created_at__range=[start_date, end_datetime]
        ).prefetch_related('stages', 'client', 'case')
        
        # Skip lawyers with no processes in the date range
        if not processes.exists():
            continue
        
        # Build lawyer name and info
        lawyer_name = f"{lawyer.first_name or ''} {lawyer.last_name or ''}".strip()
        lawyer_info = f"{lawyer_name} ({lawyer.email})"
        
        for process in processes:
            # Get latest stage
            latest_stage = process.stages.order_by('-created_at').first()
            current_stage = latest_stage.status if latest_stage else "Sin etapa"
            
            # Calculate days active
            created_date = process.created_at.date()
            days_active = (today - created_date).days
            
            # Client name
            client_name = f"{process.client.first_name or ''} {process.client.last_name or ''}".strip()
            
            # Add row to data
            data.append({
                'Abogado': lawyer_info,
                'Referencia de Proceso': process.ref,
                'Tipo de Caso': process.case.type,
                'Cliente': client_name,
                'Etapa Actual': current_stage,
                'Fecha de Creación': created_date,
                'Días Activo': days_active
            })
    
    # Create DataFrame and sort by lawyer and creation date
    df = pd.DataFrame(data)
    if not df.empty:
        df = df.sort_values(by=['Abogado', 'Fecha de Creación'])
    
    # Create Excel file
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Procesos por Abogado', index=False)
        
        # Get workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Procesos por Abogado']
        
        # Add formats
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#D9D9D9',
            'border': 1
        })
        
        # Add summary table format
        summary_format = workbook.add_format({
            'bold': True,
            'bg_color': '#F0F0F0',
            'border': 1
        })
        
        # Write headers with format
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            # Set column width based on max length in column
            if not df.empty:
                max_len = max(
                    df[value].astype(str).map(len).max(),
                    len(value)
                ) + 2
                worksheet.set_column(col_num, col_num, max_len)
        
        # Add summary table if dataframe is not empty
        if not df.empty:
            # Count processes by lawyer and create summary table
            summary_df = df.groupby('Abogado').size().reset_index(name='Total de Procesos')
            
            # Write summary table
            summary_row = len(df) + 3  # Leave 2 blank rows after main table
            worksheet.write(summary_row, 0, 'Resumen por Abogado', summary_format)
            worksheet.write(summary_row, 1, '', summary_format)
            
            summary_row += 1
            worksheet.write(summary_row, 0, 'Abogado', summary_format)
            worksheet.write(summary_row, 1, 'Total de Procesos', summary_format)
            
            for idx, row in summary_df.iterrows():
                summary_row += 1
                worksheet.write(summary_row, 0, row['Abogado'])
                worksheet.write(summary_row, 1, row['Total de Procesos'])
    
    return response


def generate_processes_by_client_report(response, start_date, end_datetime):
    """
    Generate report of processes grouped by client.
    
    Columns:
    - Cliente (User.first_name + User.last_name)
    - ID Cliente (User.identification)
    - Tipo de Documento (User.document_type)
    - Referencia de Proceso (Process.ref)
    - Tipo de Caso (Case.type)
    - Abogado Asignado (User.first_name + User.last_name)
    - Etapa Actual (Stage.status)
    - Fecha de Creación (Process.created_at)
    """
    # Filter clients if a specific user_id is provided
    if user_id:
        clients = User.objects.filter(id=user_id, role='client')
    else:
        clients = User.objects.filter(role='client')
    
    # Prepare data for Excel
    data = []
    
    for client in clients:
        # Get all processes for this client
        processes = Process.objects.filter(
            client=client,
            created_at__range=[start_date, end_datetime]
        ).prefetch_related('stages', 'lawyer', 'case')
        
        # Skip clients with no processes in the date range
        if not processes.exists():
            continue
        
        # Build client name and info
        client_name = f"{client.first_name or ''} {client.last_name or ''}".strip()
        client_id = client.identification or "No disponible"
        doc_type = client.document_type or "No especificado"
        
        for process in processes:
            # Get latest stage
            latest_stage = process.stages.order_by('-created_at').first()
            current_stage = latest_stage.status if latest_stage else "Sin etapa"
            
            # Get lawyer name
            lawyer_name = f"{process.lawyer.first_name or ''} {process.lawyer.last_name or ''}".strip()
            
            # Add row to data
            data.append({
                'Cliente': client_name,
                'ID Cliente': client_id,
                'Tipo de Documento': doc_type,
                'Referencia de Proceso': process.ref,
                'Tipo de Caso': process.case.type,
                'Abogado Asignado': lawyer_name,
                'Etapa Actual': current_stage,
                'Fecha de Creación': process.created_at.date()
            })
    
    # Create DataFrame and sort by client and creation date
    df = pd.DataFrame(data)
    if not df.empty:
        df = df.sort_values(by=['Cliente', 'Fecha de Creación'])
    
    # Create Excel file
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Procesos por Cliente', index=False)
        
        # Get workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Procesos por Cliente']
        
        # Add formats
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#D9D9D9',
            'border': 1
        })
        
        # Add summary table format
        summary_format = workbook.add_format({
            'bold': True,
            'bg_color': '#F0F0F0',
            'border': 1
        })
        
        # Write headers with format
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            # Set column width based on max length in column
            if not df.empty:
                max_len = max(
                    df[value].astype(str).map(len).max(),
                    len(value)
                ) + 2
                worksheet.set_column(col_num, col_num, max_len)
        
        # Add summary table if dataframe is not empty
        if not df.empty:
            # Count processes by client and create summary table
            summary_df = df.groupby(['Cliente', 'ID Cliente', 'Tipo de Documento']).size().reset_index(name='Total de Procesos')
            
            # Write summary table
            summary_row = len(df) + 3  # Leave 2 blank rows after main table
            worksheet.write(summary_row, 0, 'Resumen por Cliente', summary_format)
            worksheet.write(summary_row, 1, '', summary_format)
            worksheet.write(summary_row, 2, '', summary_format)
            worksheet.write(summary_row, 3, '', summary_format)
            
            summary_row += 1
            worksheet.write(summary_row, 0, 'Cliente', summary_format)
            worksheet.write(summary_row, 1, 'ID Cliente', summary_format)
            worksheet.write(summary_row, 2, 'Tipo de Documento', summary_format)
            worksheet.write(summary_row, 3, 'Total de Procesos', summary_format)
            
            for idx, row in summary_df.iterrows():
                summary_row += 1
                worksheet.write(summary_row, 0, row['Cliente'])
                worksheet.write(summary_row, 1, row['ID Cliente'])
                worksheet.write(summary_row, 2, row['Tipo de Documento'])
                worksheet.write(summary_row, 3, row['Total de Procesos'])
    
    return response


def generate_process_stages_report(response, start_date, end_datetime):
    """
    Generate report of process stages with duration information.
    
    Columns:
    - Referencia de Proceso (Process.ref)
    - Etapa (Stage.status)
    - Fecha de Inicio de Etapa (Stage.created_at)
    - Duración en Días (calculado)
    - Abogado Responsable (User.first_name + User.last_name)
    """
    # Base query to get processes in the date range
    processes_query = Process.objects.filter(
        created_at__range=[start_date, end_datetime]
    )
    
    # Prefetch related objects to optimize queries
    processes = processes_query.prefetch_related('stages', 'lawyer')
    
    # Prepare data for Excel
    data = []
    today = datetime.date.today()
    
    for process in processes:
        # Skip processes with no stages
        if not process.stages.exists():
            continue
        
        # Get all stages for the process ordered by created_at
        stages = process.stages.all().order_by('created_at')
        lawyer_name = f"{process.lawyer.first_name or ''} {process.lawyer.last_name or ''}".strip()
        
        # Calculate stage durations
        for i, stage in enumerate(stages):
            # Start date of the stage
            stage_start_date = stage.created_at.date()
            
            # End date is the start date of the next stage or today
            if i < len(stages) - 1:
                next_stage = stages[i + 1]
                stage_end_date = next_stage.created_at.date()
                is_current = False
            else:
                stage_end_date = today
                is_current = True
            
            # Calculate duration
            duration_days = (stage_end_date - stage_start_date).days
            
            # Add to data
            data.append({
                'Referencia de Proceso': process.ref,
                'Etapa': stage.status,
                'Fecha de Inicio': stage_start_date,
                'Fecha de Fin': stage_end_date if not is_current else "Actual",
                'Duración en Días': duration_days,
                'Abogado Responsable': lawyer_name,
                'Etapa Actual': is_current
            })
    
    # Create DataFrame and sort by process reference and stage start date
    df = pd.DataFrame(data)
    if not df.empty:
        df = df.sort_values(by=['Referencia de Proceso', 'Fecha de Inicio'])
    
    # Create Excel file
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Etapas de Procesos', index=False)
        
        # Get workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Etapas de Procesos']
        
        # Add formats
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#D9D9D9',
            'border': 1
        })
        
        current_stage_format = workbook.add_format({
            'bg_color': '#E8F4FF',  # Light blue background
        })
        
        # Write headers with format
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            # Set column width based on max length in column
            if not df.empty:
                max_len = max(
                    df[value].astype(str).map(len).max(),
                    len(value)
                ) + 2
                worksheet.set_column(col_num, col_num, max_len)
        
        # Apply conditional formatting for current stages
        if not df.empty:
            for row_num, row in enumerate(df.values, start=1):
                # Check if this is a current stage (last column is True)
                if row[-1]:  # Last column is 'Etapa Actual'
                    for col_num in range(len(row) - 1):  # Exclude the 'Etapa Actual' column
                        worksheet.write(row_num, col_num, row[col_num], current_stage_format)
                else:
                    for col_num in range(len(row) - 1):
                        worksheet.write(row_num, col_num, row[col_num])
            
            # Hide the 'Etapa Actual' column as it's just used for formatting
            worksheet.set_column(len(df.columns) - 1, len(df.columns) - 1, None, None, {'hidden': True})
        
        # Add summary statistics
        if not df.empty:
            # Calculate average duration by stage type
            avg_duration_by_stage = df.groupby('Etapa')['Duración en Días'].mean().reset_index()
            avg_duration_by_stage.columns = ['Etapa', 'Duración Promedio (días)']
            
            # Write summary statistics
            summary_row = len(df) + 3  # Leave 2 blank rows after main table
            worksheet.write(summary_row, 0, 'Duración Promedio por Etapa', header_format)
            
            summary_row += 1
            worksheet.write(summary_row, 0, 'Etapa', header_format)
            worksheet.write(summary_row, 1, 'Duración Promedio (días)', header_format)
            
            for idx, row in avg_duration_by_stage.iterrows():
                summary_row += 1
                worksheet.write(summary_row, 0, row['Etapa'])
                worksheet.write(summary_row, 1, round(row['Duración Promedio (días)'], 1))
    
    return response


def generate_registered_users_report(response, start_date, end_datetime):
    """
    Generate report of registered users with detailed information.
    
    Columns:
    - Email (User.email)
    - Nombre (User.first_name)
    - Apellido (User.last_name)
    - Rol (User.role)
    - Tipo de Documento (User.document_type)
    - Identificación (User.identification)
    - Contacto (User.contact)
    - Perfil Completo (User.is_profile_completed)
    - Fecha de Registro (User.created_at)
    """
    # Get users created in the date range
    users = User.objects.filter(
        created_at__range=[start_date, end_datetime]
    )
    
    # Prepare data for Excel
    data = []
    
    for user in users:
        # Map role display names
        role_display = "Cliente" if user.role == 'client' else "Abogado" if user.role == 'lawyer' else user.role
        
        # Format profile completion status
        profile_status = "Completo" if user.is_profile_completed else "Incompleto"
        
        # Add user data to the list
        data.append({
            'Email': user.email,
            'Nombre': user.first_name or "",
            'Apellido': user.last_name or "",
            'Rol': role_display,
            'Tipo de Documento': user.document_type or "No especificado",
            'Identificación': user.identification or "No disponible",
            'Contacto': user.contact or "No disponible",
            'Perfil Completo': profile_status,
            'Fecha de Registro': user.created_at.date()
        })
    
    # Create DataFrame and sort by registration date
    df = pd.DataFrame(data)
    if not df.empty:
        df = df.sort_values(by=['Fecha de Registro', 'Email'])
    
    # Create Excel file
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Usuarios Registrados', index=False)
        
        # Get workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Usuarios Registrados']
        
        # Add formats
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#D9D9D9',
            'border': 1
        })
        
        # Add conditional formats
        lawyer_format = workbook.add_format({'bg_color': '#E8F4FF'})  # Light blue for lawyers
        client_format = workbook.add_format({'bg_color': '#F2F2F2'})  # Light gray for clients
        
        # Write headers with format
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            # Set column width based on max length in column
            if not df.empty:
                max_len = max(
                    df[value].astype(str).map(len).max(),
                    len(value)
                ) + 2
                worksheet.set_column(col_num, col_num, max_len)
        
        # Apply conditional formatting based on user role
        if not df.empty:
            # Get the index of the 'Rol' column
            role_col_idx = df.columns.get_loc('Rol')
            
            for row_num, row in enumerate(df.values, start=1):
                role = row[role_col_idx]
                row_format = lawyer_format if role == "Abogado" else client_format
                
                for col_num, value in enumerate(row):
                    worksheet.write(row_num, col_num, value, row_format)
        
        # Add summary statistics
        if not df.empty:
            # Count users by role
            role_counts = df['Rol'].value_counts().reset_index()
            role_counts.columns = ['Rol', 'Cantidad']
            
            # Count complete vs incomplete profiles
            profile_counts = df['Perfil Completo'].value_counts().reset_index()
            profile_counts.columns = ['Estado', 'Cantidad']
            
            # Write user role summary
            summary_row = len(df) + 3  # Leave 2 blank rows after main table
            worksheet.write(summary_row, 0, 'Resumen por Rol de Usuario', header_format)
            
            summary_row += 1
            worksheet.write(summary_row, 0, 'Rol', header_format)
            worksheet.write(summary_row, 1, 'Cantidad', header_format)
            
            for idx, row in role_counts.iterrows():
                summary_row += 1
                worksheet.write(summary_row, 0, row['Rol'])
                worksheet.write(summary_row, 1, row['Cantidad'])
            
            # Write profile completion summary
            summary_row += 3  # Add some space
            worksheet.write(summary_row, 0, 'Resumen por Estado de Perfil', header_format)
            
            summary_row += 1
            worksheet.write(summary_row, 0, 'Estado', header_format)
            worksheet.write(summary_row, 1, 'Cantidad', header_format)
            
            for idx, row in profile_counts.iterrows():
                summary_row += 1
                worksheet.write(summary_row, 0, row['Estado'])
                worksheet.write(summary_row, 1, row['Cantidad'])
    
    return response


def generate_user_activity_report(response, start_date, end_datetime):
    """
    Generate report of user activity with detailed information.
    
    Columns:
    - Usuario (User.email)
    - Nombre (User.first_name + User.last_name)
    - Tipo de Acción (ActivityFeed.action_type)
    - Descripción de Acción (ActivityFeed.description)
    - Fecha de Acción (ActivityFeed.created_at)
    """
    # Get activities created in the date range
    activities = ActivityFeed.objects.filter(
        created_at__range=[start_date, end_datetime]
    ).select_related('user')
    
    # Prepare data for Excel
    data = []
    
    for activity in activities:
        # Build user name
        user_name = f"{activity.user.first_name or ''} {activity.user.last_name or ''}".strip()
        
        # Get readable action type
        action_types = dict(ActivityFeed.ACTION_TYPE_CHOICES)
        action_type_display = action_types.get(activity.action_type, activity.action_type)
        
        # Add activity data to the list
        data.append({
            'Usuario': activity.user.email,
            'Nombre': user_name,
            'Tipo de Acción': action_type_display,
            'Descripción': activity.description,
            'Fecha de Acción': activity.created_at
        })
    
    # Create DataFrame and sort by timestamp (newest first)
    df = pd.DataFrame(data)
    if not df.empty:
        df = df.sort_values(by=['Fecha de Acción'], ascending=False)
    
    # Create Excel file
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Actividad de Usuarios', index=False)
        
        # Get workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Actividad de Usuarios']
        
        # Add formats
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#D9D9D9',
            'border': 1
        })
        
        # Add action type formats for different actions
        create_format = workbook.add_format({'bg_color': '#E6F4EA'})  # Light green
        edit_format = workbook.add_format({'bg_color': '#E8F4FF'})    # Light blue
        delete_format = workbook.add_format({'bg_color': '#FCE8E6'})  # Light red
        other_format = workbook.add_format({'bg_color': '#F9F9F9'})   # Light gray
        
        # Write headers with format
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            # Set column width based on max length in column
            if not df.empty:
                # Use a larger width for description column
                if value == 'Descripción':
                    worksheet.set_column(col_num, col_num, 50)
                else:
                    max_len = max(
                        df[value].astype(str).map(len).max(),
                        len(value)
                    ) + 2
                    worksheet.set_column(col_num, col_num, max_len)
        
        # Apply conditional formatting based on action type
        if not df.empty:
            # Get the index of the 'Tipo de Acción' column
            action_col_idx = df.columns.get_loc('Tipo de Acción')
            
            for row_num, row in enumerate(df.values, start=1):
                action_type = row[action_col_idx]
                
                # Select format based on action type
                if 'Create' in action_type or 'create' in action_type:
                    row_format = create_format
                elif 'Edit' in action_type or 'edit' in action_type or 'Update' in action_type or 'update' in action_type:
                    row_format = edit_format
                elif 'Delete' in action_type or 'delete' in action_type:
                    row_format = delete_format
                else:
                    row_format = other_format
                
                # Apply format to all cells in the row
                for col_num, value in enumerate(row):
                    worksheet.write(row_num, col_num, value, row_format)
        
        # Add summary statistics
        if not df.empty:
            # Count activities by action type
            if 'Tipo de Acción' in df.columns:
                action_counts = df['Tipo de Acción'].value_counts().reset_index()
                action_counts.columns = ['Tipo de Acción', 'Cantidad']
                
                # Count activities by user
                user_counts = df['Usuario'].value_counts().reset_index()
                user_counts.columns = ['Usuario', 'Cantidad de Acciones']
                
                # Write action type summary
                summary_row = len(df) + 3  # Leave 2 blank rows after main table
                worksheet.write(summary_row, 0, 'Resumen por Tipo de Acción', header_format)
                
                summary_row += 1
                worksheet.write(summary_row, 0, 'Tipo de Acción', header_format)
                worksheet.write(summary_row, 1, 'Cantidad', header_format)
                
                for idx, row in action_counts.iterrows():
                    summary_row += 1
                    worksheet.write(summary_row, 0, row['Tipo de Acción'])
                    worksheet.write(summary_row, 1, row['Cantidad'])
                
                # Write user activity summary
                summary_row += 3  # Add some space
                worksheet.write(summary_row, 0, 'Resumen por Usuario (Top 10)', header_format)
                
                summary_row += 1
                worksheet.write(summary_row, 0, 'Usuario', header_format)
                worksheet.write(summary_row, 1, 'Cantidad de Acciones', header_format)
                
                # Show only top 10 most active users
                for idx, row in user_counts.head(10).iterrows():
                    summary_row += 1
                    worksheet.write(summary_row, 0, row['Usuario'])
                    worksheet.write(summary_row, 1, row['Cantidad de Acciones'])
    
    return response


def generate_lawyers_workload_report(response, start_date, end_datetime):
    """
    Generate report of lawyers and their workload.
    
    Columns:
    - Abogado (User.email + User.first_name + User.last_name)
    - Total de Procesos Asignados (conteo)
    - Procesos Activos (conteo con filtro)
    - Procesos por Tipo de Caso (desglose)
    - Fecha de Registro (User.created_at)
    """
    # Filter lawyers - either a specific lawyer or all lawyers
    if user_id:
        lawyers = User.objects.filter(id=user_id, role='lawyer')
    else:
        lawyers = User.objects.filter(role='lawyer')
    
    # Get all case types for analysis
    case_types = Case.objects.all()
    
    # Prepare data for Excel
    data = []
    
    for lawyer in lawyers:
        # Get all processes assigned to this lawyer
        all_processes = Process.objects.filter(
            lawyer=lawyer,
            created_at__range=[start_date, end_datetime]
        ).prefetch_related('stages', 'case')
        
        # Count total assigned processes
        total_processes = all_processes.count()
        
        # Skip lawyers with no processes
        if total_processes == 0:
            continue
        
        # Count active processes (processes without "Fallo" stage)
        active_processes = 0
        completed_processes = 0
        
        # Count processes by case type
        case_type_counts = {case.type: 0 for case in case_types}
        
        for process in all_processes:
            # Check if process is completed (has a "Fallo" stage)
            has_fallo = process.stages.filter(status='Fallo').exists()
            
            if has_fallo:
                completed_processes += 1
            else:
                active_processes += 1
            
            # Increment count for the process's case type
            if process.case:
                case_type_counts[process.case.type] = case_type_counts.get(process.case.type, 0) + 1
        
        # Format case type distribution
        case_type_distribution = ', '.join([
            f"{case_type}: {count}" for case_type, count in case_type_counts.items() if count > 0
        ])
        
        # Format lawyer name and info
        lawyer_name = f"{lawyer.first_name or ''} {lawyer.last_name or ''}".strip()
        lawyer_email = lawyer.email
        
        # Add lawyer data to the list
        data.append({
            'Email': lawyer_email,
            'Nombre': lawyer_name,
            'Total de Procesos Asignados': total_processes,
            'Procesos Activos': active_processes,
            'Procesos Completados': completed_processes,
            'Distribución por Tipo de Caso': case_type_distribution,
            'Fecha de Registro': lawyer.created_at.date()
        })
    
    # Create DataFrame and sort by total processes (descending)
    df = pd.DataFrame(data)
    if not df.empty:
        df = df.sort_values(by=['Total de Procesos Asignados'], ascending=False)
    
    # Create Excel file
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Carga de Trabajo de Abogados', index=False)
        
        # Get workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Carga de Trabajo de Abogados']
        
        # Add formats
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#D9D9D9',
            'border': 1
        })
        
        # Write headers with format
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            # Set column width based on max length in column
            if not df.empty:
                # Use a larger width for distribution column
                if value == 'Distribución por Tipo de Caso':
                    worksheet.set_column(col_num, col_num, 40)
                else:
                    max_len = max(
                        df[value].astype(str).map(len).max(),
                        len(value)
                    ) + 2
                    worksheet.set_column(col_num, col_num, max_len)
        
        # Add workload chart if there's data
        if not df.empty and len(df) > 1:  # Only create chart if there are multiple lawyers
            # Create a new worksheet for the chart
            chart_sheet = workbook.add_worksheet('Gráficos')
            
            # Write data for the chart
            chart_sheet.write(0, 0, 'Abogado', header_format)
            chart_sheet.write(0, 1, 'Procesos Activos', header_format)
            chart_sheet.write(0, 2, 'Procesos Completados', header_format)
            
            for i, row in df.iterrows():
                lawyer_name = row['Nombre'] or row['Email']
                chart_sheet.write(i+1, 0, lawyer_name)
                chart_sheet.write(i+1, 1, row['Procesos Activos'])
                chart_sheet.write(i+1, 2, row['Procesos Completados'])
            
            # Create a column chart
            chart = workbook.add_chart({'type': 'column'})
            
            # Add active and completed processes series
            chart.add_series({
                'name': 'Procesos Activos',
                'categories': ['Gráficos', 1, 0, len(df), 0],
                'values': ['Gráficos', 1, 1, len(df), 1],
                'fill': {'color': '#5B9BD5'}  # Blue for active processes
            })
            
            chart.add_series({
                'name': 'Procesos Completados',
                'categories': ['Gráficos', 1, 0, len(df), 0],
                'values': ['Gráficos', 1, 2, len(df), 2],
                'fill': {'color': '#70AD47'}  # Green for completed processes
            })
            
            # Configure chart
            chart.set_title({'name': 'Carga de Trabajo por Abogado'})
            chart.set_x_axis({'name': 'Abogado'})
            chart.set_y_axis({'name': 'Número de Procesos'})
            chart.set_style(10)
            
            # Insert chart
            chart_sheet.insert_chart('D2', chart, {'x_scale': 1.5, 'y_scale': 1.5})
            
            # Create a pie chart for total workload distribution
            pie_chart = workbook.add_chart({'type': 'pie'})
            
            # Prepare pie chart data
            chart_sheet.write(len(df)+3, 0, 'Abogado', header_format)
            chart_sheet.write(len(df)+3, 1, 'Total de Procesos', header_format)
            
            for i, row in df.iterrows():
                lawyer_name = row['Nombre'] or row['Email']
                chart_sheet.write(len(df)+4+i, 0, lawyer_name)
                chart_sheet.write(len(df)+4+i, 1, row['Total de Procesos Asignados'])
            
            # Add total processes series
            pie_chart.add_series({
                'name': 'Total de Procesos',
                'categories': ['Gráficos', len(df)+4, 0, len(df)*2+3, 0],
                'values': ['Gráficos', len(df)+4, 1, len(df)*2+3, 1],
                'data_labels': {'percentage': True, 'category': True}
            })
            
            # Configure pie chart
            pie_chart.set_title({'name': 'Distribución de Procesos entre Abogados'})
            pie_chart.set_style(10)
            
            # Insert pie chart
            chart_sheet.insert_chart('D20', pie_chart, {'x_scale': 1.5, 'y_scale': 1.5})
    
    return response


def generate_documents_by_state_report(response, start_date, end_datetime):
    """
    Generate a report of documents by state with filtering options.
    """
    from ..models import DynamicDocument, User
    import pandas as pd
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.series import DataPoint
    
    # Define colors for document states
    state_colors = {
        'Draft': 'FFFFE0',      # Light yellow
        'Published': 'E6F9FF',  # Light blue
        'Progress': 'E6FFE6',   # Light green
        'Completed': 'E6E6E6',  # Light gray
    }
    
    # State translations for the report
    state_translations = {
        'Draft': 'Borrador',
        'Published': 'Publicado',
        'Progress': 'En Progreso',
        'Completed': 'Completado',
    }
    
    # Query documents within the date range
    documents_query = DynamicDocument.objects.filter(
        models.Q(created_at__date__gte=start_date, created_at__date__lte=end_datetime.date()) |
        models.Q(updated_at__date__gte=start_date, updated_at__date__lte=end_datetime.date())
    ).select_related('created_by', 'assigned_to')
    
    # Apply user filter if specified
    if user_id:
        try:
            user = User.objects.get(id=user_id)
            if user.role == 'lawyer':
                documents_query = documents_query.filter(
                    models.Q(created_by=user) | models.Q(assigned_to=user)
                )
            elif user.role == 'client':
                documents_query = documents_query.filter(assigned_to=user)
        except User.DoesNotExist:
            return Response(
                {'error': f'User with id {user_id} does not exist'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    # Get the document data
    document_data = []
    today = timezone.now().date()
    
    for doc in documents_query:
        days_since_update = (today - doc.updated_at.date()).days
        
        creator_name = f"{doc.created_by.first_name} {doc.created_by.last_name}" if doc.created_by else "N/A"
        assignee_name = f"{doc.assigned_to.first_name} {doc.assigned_to.last_name}" if doc.assigned_to else "N/A"
        
        document_data.append({
            'title': doc.title,
            'state': doc.state,
            'state_es': state_translations.get(doc.state, doc.state),
            'created_by': creator_name,
            'assigned_to': assignee_name,
            'created_at': doc.created_at,
            'updated_at': doc.updated_at,
            'days_since_update': days_since_update
        })
    
    # Create dataframe for Excel export
    if document_data:
        df = pd.DataFrame(document_data)
        
        # Format dates
        df['created_at'] = df['created_at'].dt.strftime('%Y-%m-%d %H:%M')
        df['updated_at'] = df['updated_at'].dt.strftime('%Y-%m-%d %H:%M')
        
        # Create state counts for summary
        state_counts = df['state'].value_counts().to_dict()
        
        # Create Excel writer
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Documentos por Estado', index=False, columns=[
                'title', 'state_es', 'created_by', 'assigned_to', 
                'created_at', 'updated_at', 'days_since_update'
            ])
            
            # Get the worksheet to apply formatting
            worksheet = writer.sheets['Documentos por Estado']
            
            # Set column names in Spanish
            column_names = [
                'Título', 'Estado', 'Creado por', 'Asignado a', 
                'Fecha Creación', 'Última Actualización', 'Días desde Actualización'
            ]
            
            for idx, name in enumerate(column_names, start=1):
                cell = worksheet.cell(row=1, column=idx)
                cell.value = name
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")
            
            # Apply conditional formatting based on state
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for row_idx, row in enumerate(df.itertuples(), start=2):
                state = row.state
                color = state_colors.get(state, 'FFFFFF')  # Default white if state not found
                
                for col_idx in range(1, len(column_names) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
                    
                    # Center certain columns
                    if col_idx in [2, 5, 6, 7]:  # State, dates, and days columns
                        cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for idx, col in enumerate(df.columns, start=1):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(column_names[idx-1])
                ) + 2  # padding
                worksheet.column_dimensions[get_column_letter(idx)].width = max_length
            
            # Add summary sheet
            summary_data = {
                'Estado': [state_translations.get(state, state) for state in state_counts.keys()],
                'Cantidad': list(state_counts.values())
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Format summary sheet
            summary_sheet = writer.sheets['Resumen']
            summary_sheet.column_dimensions['A'].width = 15
            summary_sheet.column_dimensions['B'].width = 10
            
            # Format header
            for col in ['A', 'B']:
                cell = summary_sheet[f'{col}1']
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="BFEFFF", end_color="BFEFFF", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Apply colors to state rows
            for row_idx, state in enumerate(state_counts.keys(), start=2):
                color = state_colors.get(state, 'FFFFFF')
                cell_state = summary_sheet[f'A{row_idx}']
                cell_count = summary_sheet[f'B{row_idx}']
                
                for cell in [cell_state, cell_count]:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
            
            # Add a total row
            total_row = len(state_counts) + 2
            summary_sheet[f'A{total_row}'] = 'Total'
            summary_sheet[f'B{total_row}'] = sum(state_counts.values())
            
            for col in ['A', 'B']:
                cell = summary_sheet[f'{col}{total_row}']
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            
            # Add a pie chart
            pie = PieChart()
            labels = Reference(summary_sheet, min_col=1, min_row=2, max_row=total_row-1)
            data = Reference(summary_sheet, min_col=2, min_row=1, max_row=total_row-1)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Distribución de Documentos por Estado"
            
            # Customize chart slices
            slice_colors = ['FFFFE0', 'E6F9FF', 'E6FFE6', 'E6E6E6']
            for i, color in enumerate(slice_colors[:len(state_counts)]):
                slice = DataPoint(idx=i)
                slice.graphicalProperties.solidFill = color
                pie.series[0].data_points.append(slice)
            
            # Add the chart to the sheet
            summary_sheet.add_chart(pie, "D2")
        
        return response
    else:
        # Return empty Excel if no data found
        df = pd.DataFrame({
            'Mensaje': ['No se encontraron documentos para el período seleccionado.']
        })
        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Sin Datos', index=False)
        return response


def generate_received_legal_requests_report(response, start_date, end_datetime):
    """
    Generate report of received legal requests.
    
    Columns:
    - Nombre Solicitante (LegalRequest.first_name + LegalRequest.last_name)
    - Email (LegalRequest.email)
    - Tipo de Solicitud (LegalRequestType.name)
    - Disciplina Legal (LegalDiscipline.name)
    - Archivos Adjuntos (Count of LegalRequestFiles)
    - Descripción (LegalRequest.description)
    - Fecha de Solicitud (LegalRequest.created_at)
    """
    # Get legal requests created in the date range
    legal_requests = LegalRequest.objects.filter(
        created_at__range=[start_date, end_datetime]
    ).select_related('request_type', 'discipline').prefetch_related('files')
    
    # Prepare data for Excel
    data = []
    
    for request in legal_requests:
        # Build requester name
        requester_name = f"{request.first_name} {request.last_name}"
        
        # Count attached files
        file_count = request.files.count()
        
        # Add request data to the list
        data.append({
            'Nombre Solicitante': requester_name,
            'Email': request.email,
            'Tipo de Solicitud': request.request_type.name,
            'Disciplina Legal': request.discipline.name,
            'Archivos Adjuntos': file_count,
            'Descripción': request.description,
            'Fecha de Solicitud': request.created_at.date()
        })
    
    # Create DataFrame and sort by request date (newest first)
    df = pd.DataFrame(data)
    if not df.empty:
        df = df.sort_values(by=['Fecha de Solicitud'], ascending=False)
    
    # Create Excel file
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Solicitudes Recibidas', index=False)
        
        # Get workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Solicitudes Recibidas']
        
        # Add formats
        header_format = workbook.add_format({
            'bold': True,
            'text_wrap': True,
            'valign': 'top',
            'fg_color': '#D9D9D9',
            'border': 1
        })
        
        wrap_format = workbook.add_format({
            'text_wrap': True,
            'valign': 'top'
        })
        
        # Write headers with format
        for col_num, value in enumerate(df.columns.values):
            worksheet.write(0, col_num, value, header_format)
            
            # Set column width based on max length in column
            if not df.empty:
                if value == 'Descripción':
                    # Set a larger width for description column
                    worksheet.set_column(col_num, col_num, 50)
                else:
                    max_len = max(
                        df[value].astype(str).map(len).max(),
                        len(value)
                    ) + 2
                    worksheet.set_column(col_num, col_num, max_len)
        
        # Apply text wrap format to description column
        if not df.empty:
            desc_col_idx = df.columns.get_loc('Descripción')
            for row_num in range(1, len(df) + 1):
                worksheet.write(row_num, desc_col_idx, df.iloc[row_num-1]['Descripción'], wrap_format)
        
        # Add summary statistics if there's data
        if not df.empty:
            # Count requests by type
            request_type_counts = df['Tipo de Solicitud'].value_counts().reset_index()
            request_type_counts.columns = ['Tipo de Solicitud', 'Cantidad']
            
            # Count requests by discipline
            discipline_counts = df['Disciplina Legal'].value_counts().reset_index()
            discipline_counts.columns = ['Disciplina Legal', 'Cantidad']
            
            # Write request type summary
            summary_row = len(df) + 3  # Leave 2 blank rows after main table
            worksheet.write(summary_row, 0, 'Resumen por Tipo de Solicitud', header_format)
            
            summary_row += 1
            worksheet.write(summary_row, 0, 'Tipo de Solicitud', header_format)
            worksheet.write(summary_row, 1, 'Cantidad', header_format)
            
            for idx, row in request_type_counts.iterrows():
                summary_row += 1
                worksheet.write(summary_row, 0, row['Tipo de Solicitud'])
                worksheet.write(summary_row, 1, row['Cantidad'])
            
            # Write discipline summary
            summary_row += 3  # Add spacing
            worksheet.write(summary_row, 0, 'Resumen por Disciplina Legal', header_format)
            
            summary_row += 1
            worksheet.write(summary_row, 0, 'Disciplina Legal', header_format)
            worksheet.write(summary_row, 1, 'Cantidad', header_format)
            
            for idx, row in discipline_counts.iterrows():
                summary_row += 1
                worksheet.write(summary_row, 0, row['Disciplina Legal'])
                worksheet.write(summary_row, 1, row['Cantidad'])
            
            # Create a new sheet for charts
            chart_sheet = workbook.add_worksheet('Gráficos')
            
            # Create bar chart for request types
            chart_sheet.write(0, 0, 'Tipo de Solicitud')
            chart_sheet.write(0, 1, 'Cantidad')
            
            for i, row in enumerate(request_type_counts.values):
                chart_sheet.write(i+1, 0, row[0])
                chart_sheet.write(i+1, 1, row[1])
            
            type_chart = workbook.add_chart({'type': 'bar'})
            
            type_chart.add_series({
                'name': 'Solicitudes',
                'categories': ['Gráficos', 1, 0, len(request_type_counts), 0],
                'values': ['Gráficos', 1, 1, len(request_type_counts), 1],
                'fill': {'color': '#5B9BD5'}
            })
            
            type_chart.set_title({'name': 'Solicitudes por Tipo'})
            type_chart.set_x_axis({'name': 'Tipo de Solicitud'})
            type_chart.set_y_axis({'name': 'Cantidad'})
            
            chart_sheet.insert_chart('D2', type_chart, {'x_scale': 1.5, 'y_scale': 1.2})
            
            # Create pie chart for disciplines
            chart_sheet.write(len(request_type_counts) + 3, 0, 'Disciplina Legal')
            chart_sheet.write(len(request_type_counts) + 3, 1, 'Cantidad')
            
            for i, row in enumerate(discipline_counts.values):
                chart_sheet.write(len(request_type_counts) + 4 + i, 0, row[0])
                chart_sheet.write(len(request_type_counts) + 4 + i, 1, row[1])
            
            pie_chart = workbook.add_chart({'type': 'pie'})
            
            pie_chart.add_series({
                'name': 'Disciplinas',
                'categories': ['Gráficos', len(request_type_counts) + 4, 0, 
                               len(request_type_counts) + 4 + len(discipline_counts) - 1, 0],
                'values': ['Gráficos', len(request_type_counts) + 4, 1, 
                           len(request_type_counts) + 4 + len(discipline_counts) - 1, 1],
                'data_labels': {'percentage': True, 'category': True}
            })
            
            pie_chart.set_title({'name': 'Distribución por Disciplina Legal'})
            
            chart_sheet.insert_chart('D18', pie_chart, {'x_scale': 1.5, 'y_scale': 1.2})
    
    return response


def generate_requests_by_type_discipline_report(response, start_date, end_datetime):
    """
    Generate report of legal requests grouped by type and discipline.
    
    This report focuses on the analysis and distribution of legal requests based on their
    type and discipline, providing insights on the most common request types and legal areas.
    
    Sections:
    1. Análisis por Tipo de Solicitud
    2. Análisis por Disciplina Legal
    3. Matriz de Tipo de Solicitud vs Disciplina
    """
    # Get all legal requests in the date range
    legal_requests = LegalRequest.objects.filter(
        created_at__range=[start_date, end_datetime]
    ).select_related('request_type', 'discipline')
    
    # Get all request types and disciplines for the analysis
    request_types = LegalRequestType.objects.all()
    legal_disciplines = LegalDiscipline.objects.all()
    
    # Prepare data frames for different analyses
    # 1. Requests by Type
    type_data = []
    for req_type in request_types:
        count = legal_requests.filter(request_type=req_type).count()
        if count > 0:  # Only include types with requests
            type_data.append({
                'Tipo de Solicitud': req_type.name,
                'Cantidad': count,
                'Porcentaje': 0  # Will calculate below if there are requests
            })
    
    # 2. Requests by Discipline
    discipline_data = []
    for discipline in legal_disciplines:
        count = legal_requests.filter(discipline=discipline).count()
        if count > 0:  # Only include disciplines with requests
            discipline_data.append({
                'Disciplina Legal': discipline.name,
                'Cantidad': count,
                'Porcentaje': 0  # Will calculate below if there are requests
            })
    
    # 3. Matrix data (Types vs Disciplines)
    matrix_data = []
    for req_type in request_types:
        row_data = {'Tipo de Solicitud': req_type.name}
        for discipline in legal_disciplines:
            count = legal_requests.filter(
                request_type=req_type, 
                discipline=discipline
            ).count()
            row_data[discipline.name] = count
        
        # Only include row if there's at least one request for this type
        if any(row_data.get(discipline.name, 0) > 0 for discipline in legal_disciplines):
            matrix_data.append(row_data)
    
    # Calculate percentages if there are requests
    total_requests = legal_requests.count()
    if total_requests > 0:
        for item in type_data:
            item['Porcentaje'] = round((item['Cantidad'] / total_requests) * 100, 2)
        
        for item in discipline_data:
            item['Porcentaje'] = round((item['Cantidad'] / total_requests) * 100, 2)
    
    # Create DataFrames
    df_types = pd.DataFrame(type_data)
    df_disciplines = pd.DataFrame(discipline_data)
    df_matrix = pd.DataFrame(matrix_data)
    
    # Sort by quantity (descending)
    if not df_types.empty:
        df_types = df_types.sort_values(by='Cantidad', ascending=False)
    
    if not df_disciplines.empty:
        df_disciplines = df_disciplines.sort_values(by='Cantidad', ascending=False)
    
    # Create Excel file with multiple sheets
    with pd.ExcelWriter(response, engine='xlsxwriter') as writer:
        workbook = writer.book
        
        # 1. Summary sheet with main statistics
        summary_sheet = workbook.add_worksheet('Resumen')
        
        # Add title
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'align': 'center',
            'valign': 'vcenter'
        })
        summary_sheet.merge_range('A1:H1', 'Reporte de Solicitudes por Tipo y Disciplina', title_format)
        
        # Add date range info
        date_format = workbook.add_format({'bold': True})
        summary_sheet.write('A3', 'Período del reporte:', date_format)
        summary_sheet.write('B3', f'{start_date} al {end_datetime.date()}')
        
        # Add total requests info
        summary_sheet.write('A5', 'Total de solicitudes:', date_format)
        summary_sheet.write('B5', total_requests)
        
        # Add types and disciplines count
        summary_sheet.write('A6', 'Tipos de solicitud diferentes:', date_format)
        summary_sheet.write('B6', len(type_data))
        
        summary_sheet.write('A7', 'Disciplinas legales diferentes:', date_format)
        summary_sheet.write('B7', len(discipline_data))
        
        # Add mini charts if there's data
        if not df_types.empty and not df_disciplines.empty:
            # Type distribution chart
            type_chart = workbook.add_chart({'type': 'pie'})
            type_rows = min(5, len(df_types))  # Top 5 types or less
            
            # Write data for chart
            summary_sheet.write('D3', 'Tipo de Solicitud', date_format)
            summary_sheet.write('E3', 'Cantidad', date_format)
            
            for i in range(type_rows):
                summary_sheet.write(3+i+1, 3, df_types.iloc[i]['Tipo de Solicitud'])
                summary_sheet.write(3+i+1, 4, df_types.iloc[i]['Cantidad'])
            
            type_chart.add_series({
                'name': 'Distribución por Tipo',
                'categories': ['Resumen', 4, 3, 3+type_rows, 3],
                'values': ['Resumen', 4, 4, 3+type_rows, 4],
                'data_labels': {'percentage': True}
            })
            
            type_chart.set_title({'name': 'Principales Tipos de Solicitud'})
            summary_sheet.insert_chart('A10', type_chart, {'x_scale': 0.7, 'y_scale': 0.7})
            
            # Discipline distribution chart
            disc_chart = workbook.add_chart({'type': 'pie'})
            disc_rows = min(5, len(df_disciplines))  # Top 5 disciplines or less
            
            # Write data for chart
            summary_sheet.write('D10', 'Disciplina Legal', date_format)
            summary_sheet.write('E10', 'Cantidad', date_format)
            
            for i in range(disc_rows):
                summary_sheet.write(10+i+1, 3, df_disciplines.iloc[i]['Disciplina Legal'])
                summary_sheet.write(10+i+1, 4, df_disciplines.iloc[i]['Cantidad'])
            
            disc_chart.add_series({
                'name': 'Distribución por Disciplina',
                'categories': ['Resumen', 11, 3, 10+disc_rows, 3],
                'values': ['Resumen', 11, 4, 10+disc_rows, 4],
                'data_labels': {'percentage': True}
            })
            
            disc_chart.set_title({'name': 'Principales Disciplinas Legales'})
            summary_sheet.insert_chart('E10', disc_chart, {'x_scale': 0.7, 'y_scale': 0.7})
        
        # 2. Types Analysis Sheet
        if not df_types.empty:
            df_types.to_excel(writer, sheet_name='Análisis por Tipo', index=False)
            
            # Get the worksheet
            type_sheet = writer.sheets['Análisis por Tipo']
            
            # Format headers
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': '#D9D9D9',
                'border': 1
            })
            
            for col_num, value in enumerate(df_types.columns.values):
                type_sheet.write(0, col_num, value, header_format)
                # Set column widths
                if value == 'Tipo de Solicitud':
                    type_sheet.set_column(col_num, col_num, 30)
                else:
                    type_sheet.set_column(col_num, col_num, 15)
            
            # Add bar chart
            chart = workbook.add_chart({'type': 'column'})
            
            chart.add_series({
                'name': 'Cantidad',
                'categories': ['Análisis por Tipo', 1, 0, len(df_types), 0],
                'values': ['Análisis por Tipo', 1, 1, len(df_types), 1],
                'fill': {'color': '#5B9BD5'}
            })
            
            chart.set_title({'name': 'Solicitudes por Tipo'})
            chart.set_x_axis({'name': 'Tipo de Solicitud'})
            chart.set_y_axis({'name': 'Cantidad'})
            
            # Place the chart below the data
            type_sheet.insert_chart(f'A{len(df_types) + 3}', chart, {'x_scale': 1.5, 'y_scale': 1})
        
        # 3. Disciplines Analysis Sheet
        if not df_disciplines.empty:
            df_disciplines.to_excel(writer, sheet_name='Análisis por Disciplina', index=False)
            
            # Get the worksheet
            discipline_sheet = writer.sheets['Análisis por Disciplina']
            
            # Format headers
            for col_num, value in enumerate(df_disciplines.columns.values):
                discipline_sheet.write(0, col_num, value, header_format)
                # Set column widths
                if value == 'Disciplina Legal':
                    discipline_sheet.set_column(col_num, col_num, 30)
                else:
                    discipline_sheet.set_column(col_num, col_num, 15)
            
            # Add bar chart
            chart = workbook.add_chart({'type': 'column'})
            
            chart.add_series({
                'name': 'Cantidad',
                'categories': ['Análisis por Disciplina', 1, 0, len(df_disciplines), 0],
                'values': ['Análisis por Disciplina', 1, 1, len(df_disciplines), 1],
                'fill': {'color': '#70AD47'}
            })
            
            chart.set_title({'name': 'Solicitudes por Disciplina Legal'})
            chart.set_x_axis({'name': 'Disciplina Legal'})
            chart.set_y_axis({'name': 'Cantidad'})
            
            # Place the chart below the data
            discipline_sheet.insert_chart(f'A{len(df_disciplines) + 3}', chart, {'x_scale': 1.5, 'y_scale': 1})
        
        # 4. Matrix Analysis Sheet (Types vs Disciplines)
        if not df_matrix.empty:
            df_matrix.to_excel(writer, sheet_name='Matriz Tipo-Disciplina', index=False)
            
            # Get the worksheet
            matrix_sheet = writer.sheets['Matriz Tipo-Disciplina']
            
            # Set header formats
            for col_num, value in enumerate(df_matrix.columns.values):
                matrix_sheet.write(0, col_num, value, header_format)
                matrix_sheet.set_column(col_num, col_num, 15)
            
            # Add heat map formatting (conditional formatting)
            # First find the maximum value in the matrix for scaling
            numeric_cols = [col for col in df_matrix.columns if col != 'Tipo de Solicitud']
            if numeric_cols:
                max_value = df_matrix[numeric_cols].max().max()
                
                # Add a color scale (heat map)
                if max_value > 0:
                    # Apply to all cells except the first column (type names)
                    for col_idx, col in enumerate(numeric_cols, start=1):
                        # We add 1 to row_idx to account for the header row
                        for row_idx in range(len(df_matrix)):
                            cell_value = df_matrix.iloc[row_idx][col]
                            
                            # Skip zeros
                            if cell_value == 0:
                                continue
                            
                            # Calculate intensity (0-255) based on the value
                            intensity = min(255, int(180 * cell_value / max_value) + 50)
                            # Convert to hex color (from light to dark blue)
                            r = 255 - intensity
                            g = 255 - intensity
                            b = 255
                            color = f'#{r:02X}{g:02X}{b:02X}'
                            
                            # Create a format for this cell
                            cell_format = workbook.add_format({
                                'bg_color': color,
                                'align': 'center'
                            })
                            
                            # Apply the format to the cell
                            matrix_sheet.write(row_idx + 1, col_idx, cell_value, cell_format)
            
            # Add a heatmap chart if there's enough data
            if len(df_matrix) > 1 and len(numeric_cols) > 1:
                chart = workbook.add_chart({'type': 'heatmap'})
                
                chart.add_series({
                    'categories': ['Matriz Tipo-Disciplina', 0, 1, 0, len(numeric_cols)],
                    'values': ['Matriz Tipo-Disciplina', 1, 1, len(df_matrix), len(numeric_cols)],
                    'name': 'Distribución',
                    'data_labels': {'value': True}
                })
                
                chart.set_title({'name': 'Mapa de Calor: Tipo vs Disciplina'})
                chart.set_x_axis({'name': 'Disciplina Legal'})
                chart.set_y_axis({'name': 'Tipo de Solicitud'})
                
                # Place the chart below the data
                matrix_sheet.insert_chart(f'A{len(df_matrix) + 3}', chart, {'x_scale': 1.5, 'y_scale': 1.5})
        
        # 5. List of requests with type and discipline
        detailed_data = []
        for request in legal_requests:
            detailed_data.append({
                'Nombre Solicitante': f"{request.first_name} {request.last_name}",
                'Email': request.email,
                'Tipo de Solicitud': request.request_type.name,
                'Disciplina Legal': request.discipline.name,
                'Archivos Adjuntos': request.files.count(),
                'Fecha de Solicitud': request.created_at.date()
            })
        
        if detailed_data:
            df_details = pd.DataFrame(detailed_data)
            df_details = df_details.sort_values(by=['Tipo de Solicitud', 'Disciplina Legal', 'Fecha de Solicitud'])
            df_details.to_excel(writer, sheet_name='Detalle de Solicitudes', index=False)
            
            # Format the details sheet
            details_sheet = writer.sheets['Detalle de Solicitudes']
            
            for col_num, value in enumerate(df_details.columns.values):
                details_sheet.write(0, col_num, value, header_format)
                # Set column widths based on content
                max_len = max(
                    df_details[value].astype(str).map(len).max(),
                    len(value)
                ) + 2
                details_sheet.set_column(col_num, col_num, min(max_len, 30))  # Cap at 30
    
    return response